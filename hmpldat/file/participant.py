"""
object representation of a participant

Ian Zurutuza
7 August 2020
"""

# from standard library (no install required)
import glob
import os
from pathlib import Path
import sys
from pprint import pprint

# from submodel
from hmpldat.file.task import task
import hmpldat.file.rawetg
import hmpldat.file.detected
import hmpldat.align.temporal

# 3rd party (separate install required)
import pandas as pd
from openpyxl import load_workbook
import re

### Typical task order (VMIB) ###
"""
1.  TS              --  trouble shooting, 13 balls, may occur more than once a session
2.  QS_open         --  quiet standing eyes open blank screen (no start sequence)
3.  QS_closed       --  quiet standing eyes closed (no start sequence)
4.  QS_cross        --  quiet standing eyes open cross on screen (no start sequence)
5.  Fix             --  fixation, alternation between randomly appearing balls and 
                        center cross between each appearance (start sequence "Ready?, 3, 2, 1")
6.  HP              --  horizontal pursuit, ball starts center and rolls to the left 
                        or right (start sequence "Ready?, 3, 2, 1")
7.  PP              --  periferal pursuit, ball start on the left or right and rolls 
                        across the screen (start sequence "Ready?, 3, 2, 1")
8.  AP              --  angled pursuit, ball rolls across grid similarly to intercept 
                        and avoid but NO user(blue) ball (start sequence "Ready?, 3, 2, 1")
    
    ~~~~ potential to take a break, especially younger and ASD participants ~~~~

9.  BM              --  body movement, move user ball to green rectangle(safezone) 
                        (start sequence "Ready?, 3, 2, 1")
10. Int             --  intercept, block target(red/orange) ball with user(blue) ball 
                        target rolls no so "randomly" across grid (start sequence "Ready?, 3, 2, 1")
11. Avoid           --  avoid, avoid target(red/orange) ball with user(blue) ball 
                        target rolls no so "randomly" across grid (start sequence "Ready?, 3, 2, 1")
12. VT              --  visual tracking, use body weight to keep black disk(user_disk) on 
                        top of blue disk(target_disk) (Race start sequence)
13. Ducks           --  ducks, quack quack quack quack quack quack quack quack quack quack
"""

class participant:
    """represent all of a participants tasks and etg files"""

    def __init__(self, experiment, name):
        """
        I think moving from name to a experiment name combination would work better,
        Since we can split file names with _ to get relevant information and ,
        factor the experiment in our code in the future if experiments will
        have something different.
        """
        self.experiment = experiment
        self.name = name

        """
        tasks may vary across participants and studies
        tasks may repeat (typical for tasks with distractors or environment)
        this list is not ordered? Or can it be ordered?
        """

        """
        I was thinking a dictionary would be better with tasks as the keys, 
        Maybe we can create a list after we can order everything or keep a
        ordered list of just keys.
        """
        self.tasks = {}

        """
        do not create writer at this point (we may load from file)?
        """
        self.excel_writer = None

        """
        these files belong to the participant, they contain data for all tasks
        """
        self.rawetg_path = None
        self.video_path = None
        self.detections_path = None

        """
        Object detections and rawetg data should be merged before spilting by annotations
        """
        self.rawetg_and_objects_df = None

        """
        Keep track of the starts and end for each task, currently by using annotations in the rawetg data
        """
        self.task_start_and_ends = None


    def __str__(self):
        return f

    def load_file_listing(self, file_listing_path):
        """"""

        # TODO: test whether this can hadle path object or need to cast to string
        self.excel_writer = pd.ExcelWriter(file_listing_path, engine="openpyxl", mode="r")
        self.excel_writer.book = load_workbook(file_listing_path)

        try:
            self.file_listing = pd.read_excel(
                self.excel_writer,
                sheet_name=self.name,
                index_col=[0, 1, 2, 3],
                engine="openpyxl",
                converters={
                    "path": Path  # convert path string to path object (TODO: append drive|folder mount location)
                },
            )
        except KeyError:
            sys.exit(
                f"ERROR: file listing for participant: '{self.name}' does not exist.\n\tSee method: create_file_listing()\n\nEXIT"
            )

        self.file_listing = self.file_listing.reset_index()
        self.file_listing = self.file_listing.set_index(['task_name', 'task_round']).sort_index()

    def create_tasks(self,):
        task_ids = self.file_listing.index.unique() # Get the list of found tasks and rounds

        for task_id in task_ids:
            task_files=self.file_listing.loc[task_id,["file_type", "path"]]
            if "cortex" in task_files["file_type"].to_numpy():
                cortex_path = task_files.loc[task_files["file_type"] == "cortex", "path"][0]
            else:
                raise ValueError(f"Missing cortex file for task {task_id}")
            if "dflow_mc" in task_files["file_type"].to_numpy():
                mc_path = task_files.loc[task_files["file_type"] == "dflow_mc", "path"][0]
            else:
                pass
                # raise ValueError(f"Missing dflow_mc file for task {task_id}")
            if "dflow_rd" in task_files["file_type"].to_numpy():
                rd_path = task_files.loc[task_files["file_type"] == "dflow_rd", "path"][0]
            else:
                rd_path = None
            if "dflow_ev" in task_files["file_type"].to_numpy():
                ev_path = task_files.loc[task_files["file_type"] == "dflow_ev", "path"][0]
            else:
                ev_path = None
            try: 
                self.tasks[task_id] = task(
                    task_id[0], 
                    task_id[1], 
                    mc_path, 
                    cortex_path, 
                    # only a section of this aligned data is passed to the task object
                    # TODO: at the moment alignment requires the first row to be the annotation
                    # this cuts off several rows of detected grid that appears before the cross 
                    self.rawetg_and_objects_df[
                        (
                            (self.rawetg_and_objects_df['RecordingTime [ms]'] >= self.task_start_and_ends[task_id[0]]['start']) 
                            & (self.rawetg_and_objects_df['RecordingTime [ms]'] <= self.task_start_and_ends[task_id[0]]['end'])
                        )], 
                    rd_path, 
                    ev_path
                    ) 
            except Exception as e: 
                print(f"There was an error on {task_id} : {e}")
                continue 
            # self.tasks[task_id] = task(task_id[0], task_id[1], mc_path, cortex_path, rd_path, ev_path)

        # pprint(self.tasks)

    def load_rawetg_and_objects(self):
        """
        
        assuming rawetg_path and detections_path have been defined

        returns nothing, open and aligns rawetg + object detection data
        
        """

        # open rawetg 
        rawetg_df = hmpldat.file.rawetg.open(self.rawetg_path)

        # open object detections
        detections_df = hmpldat.file.detected.open(self.detections_path)

        # add (u,v) location for objects (center of detection)
        detections_df = detections_df.assign(
            u = (detections_df['left'] + detections_df['right']) / 2,
            v = (detections_df['top'] + detections_df['bottom']) / 2
        )

        # reformat (each row contain the information for one scene frame)
        detections_df = hmpldat.file.detected.reformat(detections_df)

        # correct frame_time
        # weird problem exists with the timestamps in the first few frames of the video recording
        # see fix_frame_time() for more info
        detections_df = hmpldat.file.detected.fix_frame_time(detections_df)

        # calculate frame number for each entry in rawetg
        # required if you want to align by frame_number 
        # rawetg_df['frame_number'] = hmpldat.file.rawetg.frame_number(rawetg_df, how='floor')

        # align these two frames
        self.rawetg_and_objects_df = hmpldat.align.temporal.objects_to_rawetg(detections_df, rawetg_df, how='time')

        # self.rawetg_and_objects_df = self.rawetg_and_objects_df.set_index("RecordingTime [ms]", drop=False)
        # print(self.rawetg_and_objects_df.columns)

        annotated_tasks = hmpldat.file.rawetg.get_tasks(rawetg_df)

        # TODO:  drop rows with annotation?
        
        # these column names need to change
        converters = {
            'calib': 'ts',
            'qs_eo': 'qs_open',
            'qs_eo_cross': 'qs_cross',
            'qs_ec': 'qs_closed',
            'fixation': 'fix',
            'hp': 'hp',
            'pp': 'pp',
            'ap': 'ap',
            'bm': 'bm',
            'intercept': 'int',
            'avoid': 'avoid',
            'vt': 'vt',
            'ducks': 'ducks'
        }

        annotated_tasks.columns = annotated_tasks.columns.map(converters, na_action=None)

        self.task_start_and_ends = annotated_tasks

    def find_annotationed_task(self):

        # the correct section of the rawetg+objects dataframe


        return df


    """
    The function that parses the file name , currently assumes the round is 1 if it is missing on the file name
    """
    def parse_filename(self, file_path, file_type):
        file_name = file_path.name.lower()
        file_name = file_name.split(".")[0]
        file_name = file_name.replace("-", "_")
        parts=file_name.split("_")
        experiment=parts[0]
        id=parts[1]
        matchObj = re.match(r'\d{6,6}', parts[-1]) # check if contains date at the end
        if matchObj:
            parts.pop()
        if (file_type == "cortex"):
            if "qs" in parts[2]:
                try:
                    task_round = int(parts[3][-1])
                except:
                    task_round = 1
                task_name = parts[2] + "_" + parts[3][:-1]
            elif "ducks" in parts[2]:
                try:
                    task_round = int(parts[2][-1])
                except:
                    task_round = 1
                task_name = "ducks"
            else:
                try:
                    task_round = int(parts[2][-1])
                except:
                    task_round = 1
                task_name = parts[2][:-1]
        elif (file_type == "dflow"):
            if "dflow" in parts:
                parts.remove("dflow")
            if "qs" in parts[2]:
                try:
                    task_round = int(parts[4][-4:])
                except:
                    task_round = 1
                task_name = parts[2] + "_" + parts[3]
                file_type = file_type + "_" + parts[4][:2]
            else:
                try:
                    task_round = int(parts[3][-4:])
                except:
                    task_round = 1
                task_name = parts[2]
                file_type = file_type + "_" + parts[3][:2]
        else:
            raise ValueError("unexpected file_type")
        return [task_name, task_round, file_type]


    def create_file_listing(self, file_listing_path, search_folder):
        """parse file system for task and video files"""

        print(file_listing_path, file_listing_path.exists())
        print(search_folder)

        # if this excel file already exists, append to it.
        # otherwise create it
        if file_listing_path.exists():
            self.excel_writer = pd.ExcelWriter(
                file_listing_path, engine="openpyxl", mode="a"
            )
            if self.name in load_workbook(file_listing_path, read_only=True).sheetnames:
                raise ValueError(f"Sheet '{self.name}' already exists in {file_listing_path.name}.")

        else:
            self.excel_writer = pd.ExcelWriter(
                file_listing_path, engine="openpyxl"
            )

        # open excel file check if sheet already exists for this participant

        cortex_files = []
        """
        dflow handling part , I was planing to go with fixed paths for files,
        instead of searching all files and adding the ones found.
        """

        dflow_path = search_folder / 'DFlow' / (self.experiment + "_" + self.name)
        dflow_files = [x for x in dflow_path.iterdir() if ".png" not in x.name]
        cortex_path = search_folder / 'Cortex' / (self.experiment + "_" + self.name) / 'TRC'
        cortex_files = [x for x in cortex_path.iterdir()]

        file_info = []
        for f in cortex_files:
            data = self.parse_filename(f, "cortex")
            data.append(str(f))
            file_info.append(data)

        for f in dflow_files:
            data = self.parse_filename(f, "dflow")
            data.append(str(f))
            file_info.append(data)

        file_info = pd.DataFrame(file_info,
                                 columns=["task_name", "task_round", "file_type", "path"])
        file_info = file_info.set_index(['task_name', 'task_round']).sort_index()

        """
        Need help with turning to relative path
        """
        #file_info['path'] = file_info['path'].map(lambda p: p.relative_to(MNT))

        if len(file_info) == 0:
            print(f"Error: no files found for '{self.name}' in '{search_folder}'")
            return

        file_info.to_excel(self.excel_writer, sheet_name=self.name)

        # normalize column widths
        self.excel_writer.sheets[self.name].column_dimensions['A'].width = 8
        self.excel_writer.sheets[self.name].column_dimensions['B'].width = 25
        self.excel_writer.sheets[self.name].column_dimensions['C'].width = 12
        self.excel_writer.sheets[self.name].column_dimensions['D'].width = 12
        self.excel_writer.sheets[self.name].column_dimensions['F'].width = 150

        # save and close file
        self.excel_writer.save()
        self.excel_writer.close()


    def dflow_task_order(self):
        """return order according to running dflow time"""

        # TODO: signal non continuous dflow time.
        # print('dflow running time is not continuous!')


def example_usage():
    experiment = "VMIB"
    participant_name = '005'
    # create participant object
    p = participant(experiment, participant_name)
    p.create_file_listing(
        Path("C:/UNT HTC/file_listing_VMIB.xlsx"),
        Path("C:/UNT HTC/selected_data/VMIB/Data")
    )
    # p.load_file_listing(
    #     Path("C:/UNT HTC/file_listing_VMIB.xlsx")
    # )
    # p.rawetg_path = Path('C:/UNT HTC/selected_data/VMIB/Data/ETG/Metrics Export/VMIB_047_RawETG.txt')
    # p.detections_path = Path('C:/UNT HTC/selected_data/VMIB/Data/detections/vmib_047-1-unpack.txt')
    # output_data_dir = Path(f'C:/UNT HTC/{p.experiment}/{p.name}')
    # if output_data_dir.exists():
    #     i = input("directory already exists. continue? [Y/n] ")
    #     if i.lower() == 'n':
    #         exit()
    # else:
    #     output_data_dir.mkdir(parents=True)
    #
    # # p.video_path =
    # p.load_rawetg_and_objects()
    # p.create_tasks()
    #
    # task = p.tasks[("int",1)]
    # print(task.name, hasattr(task, 'merged'))
    #
    # # TODO: keep n sample per task -> one camera model per task
    # # samples =
    #
    # # skip tasks without merged data
    # if not hasattr(task, 'merged'):
    #     pass
    #
    # task.merged.to_csv(output_data_dir / ("_".join([task.name, str(task.trial)]) + '.csv'))
    #
    # required_columns = task.merged.loc[:, task.merged.columns.str.contains(
    #     '.head|target(\.|v)|cross(\.|v)|user(\.|v)|safezone|grid')].columns
    # cleaned = task.merged.dropna(subset=required_columns, thresh=29)
    #
    # cleaned.to_csv(output_data_dir / ("_".join([task.name, str(task.trial), 'cleaned']) + '.csv'))
    # print(task.horizon_height())


    # print(p.file_listing)
    #print(p.file_listing[p.file_listing['task_name'] == 'avoid'])

if __name__ == "__main__":
    example_usage()
